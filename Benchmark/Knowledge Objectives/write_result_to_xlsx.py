import os
import json
import pandas as pd
import glob
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
from matplotlib import pyplot as plt
from matplotlib.ticker import MaxNLocator
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import json
working_directory = "/home/lym/GraphRAG4OralHealth/Benchmark/Knowledge Objectives/Answer"
#获得目录下所有模型名称
def get_model_names(directory_path):
    model_names = []
    for folder_name in os.listdir(directory_path):
        folder_path = os.path.join(directory_path, folder_name)
        if os.path.isdir(folder_path):
            model_names.append(folder_name)
    return model_names
model_list = get_model_names(working_directory)
topic_list=["MedicalHumanity","Clinical","Dentistry","Medical"]
result_dict = {}
for topic in topic_list:
    result_dict[topic] = {}
    json_files = glob.glob(os.path.join("./", topic, "*.json"))
    for file in json_files:
        with open(file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
            name=os.path.basename(file).split("_")[0]
            result_dict[topic][name] = []
            for question in json_data.get('独立题', []):
                ques=question['题干'] if '题干' in question.keys() else question['题目']
                ans=question['答案']
                result_dict[topic][name].append({
                    "题目": ques,
                    "选项": question['选项'],
                    "答案": ans
                })
            for question in json_data.get('共用题干题', []):
                for idx in range(len(question['答案'])):
                    ques=question['题干'][idx] if '题干' in question.keys() else question['题目'][idx]
                    ans=question['答案'][idx]
                    result_dict[topic][name].append({
                        "题目": ques,
                        "选项": question['选项'][idx],
                        "答案": ans
                    })
                    

            for question in json_data.get('共用备选题', []):
                for idx in range(len(question['答案'])):
                    ques=question['题干'][idx] if '题干' in question.keys() else question['题目'][idx]
                    ans=question['答案'][idx]
                    result_dict[topic][name].append({
                        "题目": ques,
                        "选项": question['选项'],
                        "答案": ans
                    })
import os
import json

def process_json_files(base_dir):
    """
    逐层遍历目录结构，读取所有 JSON 文件，并组织为结构化数据
    返回格式：model_result[模型名][科目名] = [题目列表]
    """
    model_result = {}

    # 遍历模型目录
    for topic in os.listdir(base_dir):
        topic_path = os.path.join(base_dir, topic)
        if not os.path.isdir(topic_path):
            continue
        model_result[topic] = {}
        # 遍历科目目录
        for subject in os.listdir(topic_path):
            subject_path = os.path.join(topic_path, subject)
            subject1=subject.split(".j")[0].split("_")[-2]
            model_result[topic][subject1] = []
            with open(subject_path, 'r', encoding='utf-8') as f:
                try:
                    data = json.load(f)
                    for question in data:
                        ques = question.get('题干') or question.get('题目', '')
                        option = question.get('选项', [])
                        ans = question.get('prediction', '')

                        model_result[topic][subject1].append({
                            "题目": ques,
                            "选项": option,
                            "答案": ans
                        })
                except json.JSONDecodeError as e:
                    print(f"JSON解析错误：{subject_path} - {e}")

    return model_result

def eval_model(model_result, result_dict):
    """模型评估"""
    model_eval = {}
    for topic in topic_list:
        model_eval[topic] = {}
        for subject in result_dict[topic].keys():
            model_eval[topic][subject] = {}
            model_eval[topic][subject]["正确"] = []
            model_eval[topic][subject]["错误"] = []
            if result_dict[topic].keys()!= model_result[topic].keys():
                print("result",result_dict[topic].keys())
                print("model",model_result[topic].keys())
            for gt,pred in zip(result_dict[topic][subject],model_result[topic][subject]):
                # 判断正确性
                if gt["答案"] == pred['答案'] or pred['答案'].startswith(gt["答案"]): 
                    model_eval[topic][subject]["正确"].append(pred)
                else:
                    model_eval[topic][subject]["错误"].append(pred)
            # 计算准确率
            correct_count = len(model_eval[topic][subject]["正确"])
            total_count = len(result_dict[topic][subject])
            accuracy = correct_count / total_count if total_count > 0 else 0
            model_eval[topic][subject]["准确率"] = accuracy
            model_eval[topic][subject]["总题数"] = total_count
            model_eval[topic][subject]["正确数"] = correct_count
    return model_eval
model_result_dict = {}
for model_name in model_list:
    model_result_dict[model_name] = {}
    base_dir = '/home/lym/GraphRAG4OralHealth/Benchmark/Knowledge Objectives/Answer/' + model_name
    # 遍历目录下的所有JSON文件并生成请求
    model_result = process_json_files(base_dir)
    print(f"模型名称: {model_name}")
    model_eval=eval_model(model_result, result_dict)
    model_result_dict[model_name] = model_eval

# with open("model_result_dict.json", "w", encoding="utf-8") as f:
#     json.dump(model_result_dict, f, ensure_ascii=False, indent=4)
# 构建树状结构，方便汇总
subject_tree = defaultdict(lambda: defaultdict(dict))

# 填充结构 subject_tree[模型][topic][subject] = {准确率、总题数、正确数}
for model_name, topic_data in model_result_dict.items():
    for topic, subject_data in topic_data.items():
        for subject, values in subject_data.items():
            subject_tree[model_name][topic][subject] = {
                "正确数": values["正确数"],
                "总题数": values["总题数"],
                "准确率": values["准确率"]
            }

# 构建完整的 subject 层次结构
from collections import OrderedDict

# 手动定义一个层次结构（可以自动构建树，如果你有ontology或关系图）
from collections import OrderedDict

from collections import OrderedDict

subject_hierarchy = OrderedDict({
    "客观题": OrderedDict({
        "口腔专业知识": OrderedDict({
            "口腔基础医学": [
                "口腔组织病理学",
                "口腔解剖生理学",
                "牙体牙髓病学",
                "牙周病学",
                "儿童口腔医学",
                "口腔黏膜病医学",
                "口腔颌面外科学",
                "口腔修复学",
                "口腔颌面医学影像诊断学",
                "口腔预防医学"
            ]
        }),
        "医学专业知识": OrderedDict({
            "基础医学": [
                "生物化学",
                "医学微生物学",
                "医学免疫学",
                "药理学"
            ],
            "临床医学": [
                "诊断学",
                "内科学",
                "外科学",
                "妇产科学",
                "儿科学",
                "预防医学"
            ]
        }),
        "医学人文综合": OrderedDict({})
    })
})



# ------- 重新构建数据 DataFrame，适配你的格式 --------
def flatten_with_count(hierarchy, subject_tree, model_list):
    def dfs(node, indent=""):
        local_rows = []
        print(f"node: {node}")
        if node==OrderedDict():
                full_name = f"{indent}医学人文综合"
                clean_parent = "医学人文综合"
                subject_total = 0
                subject_acc = {}
                for model in model_list:
                    for topic in subject_tree[model].keys():
                        if clean_parent in subject_tree[model][topic]:
                            info = subject_tree[model][topic][clean_parent]
                            subject_total = info["总题数"]
                            subject_acc[model] = round(info["准确率"], 4)
                            break

                local_rows.append((full_name, False, subject_total, subject_acc))
        elif isinstance(node, dict) and node!=OrderedDict():
            for parent, children in node.items():
                full_name = f"{indent}{parent}"
                clean_parent = parent.strip()

                children_data = dfs(children, indent + "  ")

                total = sum(item[2] for item in children_data if not item[1])  # 只加叶子节点（is_parent=False）
                model_acc = {}
                for model in model_list:
                    acc_list = [(item[3][model], item[2]) for item in children_data if model in item[3] and not item[1]]
                    if acc_list:
                        weighted = sum(a * c for a, c in acc_list)
                        count_sum = sum(c for _, c in acc_list)
                        model_acc[model] = round(weighted / count_sum, 4)
                print(f"parent: {parent}, subject_total: {total}, model_acc: {model_acc}")
                local_rows.append((full_name, True, total, model_acc))
                local_rows.extend(children_data)
            
        elif isinstance(node, list):
            for leaf in node:
                full_name = f"{indent}{leaf}"
                clean_leaf = leaf.strip()
                subject_total = 0
                subject_acc = {}
                for model in model_list:
                    for topic in subject_tree[model].keys():
                        if clean_leaf in subject_tree[model][topic]:
                            info = subject_tree[model][topic][clean_leaf]
                            subject_total = info["总题数"]
                            subject_acc[model] = round(info["准确率"], 4)
                            break
                print(f"leaf: {leaf}, subject_total: {subject_total}, subject_acc: {subject_acc}")
                local_rows.append((full_name, False, subject_total, subject_acc))

        return local_rows

    return dfs(hierarchy)



flat_rows = flatten_with_count(subject_hierarchy, subject_tree, model_list)[:-1]
print(f"📊 扁平化数据行数: {len(flat_rows)}")
# -------- Excel 写入 + 样式增强 --------
from openpyxl import load_workbook
from openpyxl.styles import Font

excel_path = "模型评估结果_最终版.xlsx"
custom_model_order = [
    "GPT-4-turbo", "GPT-3.5-turbo", "deepseek-v3", "deepseek-r1",
    "Qwen-max", "Qwen-plus","QWen3-14B", "QWen2.5-7b-instruct", "QWen3-8B","ChatGLM-4",
    "XunfeiSpark", "DentalMind_base", "DentalMind_o1", "DentalMind_graph","DeepSeek-R1-Distill-Qwen-7B","QWen25-Math-7B","DeepSeek-R1-Distill-Qwen-1.5B","QWen25-14B","DeepSeek-R1-Distill-Qwen-14B","DentalMind_graph_o1"
]

columns = ["科目", "数量"] + custom_model_order
table_data = []
for name, is_parent, count, acc_dict in flat_rows:
    row = {"科目": name, "数量": count}
    for model in custom_model_order:
        row[model] = acc_dict.get(model, "")
    table_data.append(row)

df = pd.DataFrame(table_data)[columns]  # 明确列顺序

# 写入 Excel
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name="准确率统计")

    # 样式
    workbook = writer.book
    worksheet = writer.sheets["准确率统计"]
    for row_idx, (name, is_parent, *_rest) in enumerate(flat_rows, start=2):
        if is_parent:
            cell = worksheet[f"A{row_idx}"]
            cell.font = Font(bold=True)

print(f"✅ Excel 写入完成，列顺序为自定义模型顺序：{custom_model_order}")

