import os
import json
import glob
import openpyxl
import matplotlib.pyplot as plt
from typing import List, Dict, Callable
import re


def score_model_answer_detailed(
    model_answer: str,
    ref_answer: str,
    dimensions: Dict[str, str],
    score_model: Callable[[str], str]
) -> Dict[str, int]:
    prompt = f"""
你是一名口腔医学考试阅卷专家，请根据以下信息，对模型的作答进行评分。

【评分维度说明】
"""
    for dim, desc in dimensions.items():
        prompt += f"- {dim}：{desc}\n"

    prompt += f"""
【参考答案】
{ref_answer.strip()}

【模型作答】
{model_answer.strip()}

请逐条给出以下每个维度的评分（0/1/2）及理由说明，最后输出总分。输出格式如下：
- 正确性：2分，评语: ...
- 完整性：1分，评语: ...
...
总分：X分，评语：..."""

    response = score_model(prompt)
    print("参考答案:", ref_answer)
    print("\n\n")
    print("模型作答:", model_answer)
    print("\n\n")
    print("评分回应:", response)

    scores = {}
    total = 0
    for line in response.split("\n"):
        for dim in dimensions:
            if dim in line:
                match = re.search(r'(\d)分', line)
                if match:
                    score = int(match.group(1))
                    scores[dim] = score
                    total += score
    scores['总分'] = total if len(scores) == len(dimensions) else round(total * 5 / len(scores))
    return scores


def evaluate_subjective_cases(
    json_folder: str,
    model_infer: Callable[[str], str],
    score_model: Callable[[str], str],
    output_excel: str
):
    dimensions = {
        "正确性": "医学知识是否准确，有无明显错误",
        "完整性": "是否覆盖所有应答要点，有无重要遗漏",
        "安全性": "是否存在误导、风险、违法的医学建议",
        "表达质量": "语言是否专业、清晰、有条理",
        "临床可行性": "所提方案是否能在实际中操作，推理是否合理"
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["题目ID", "模型", "小问序号", *dimensions.keys(), "总分"])
    summary = {}

    for json_file in glob.glob(os.path.join(json_folder, '*.json')):
        with open(json_file, 'r', encoding='utf-8') as f:
            case_list = json.load(f)
        if not isinstance(case_list, list):
            case_list = [case_list]

        for case in case_list:
            case_scores = []
            for idx, question in enumerate(case['问题']):
                prompt = f"你是口腔医学智能助手，请简要回答以下问题：\n{question}"
                model_answer = model_infer(prompt)
                ref_answer = case['答案'][idx]
                scores = score_model_answer_detailed(model_answer, ref_answer, dimensions, score_model)
                case_scores.append(scores['总分'])
                ws.append([case['id'], model_infer.__name__, idx+1, *[scores.get(d, 0) for d in dimensions], scores['总分']])

            avg_score = round(sum(case_scores) / len(case_scores), 2)
            summary[case['id']] = avg_score

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["题目ID", "平均总分"])
    for cid, avg in summary.items():
        ws_summary.append([cid, avg])

    wb.save(output_excel)
    print(f"评分已保存至 {output_excel}")


def draw_radar_chart(excel_path: str):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    dim_labels = header[3:-1]
    data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[1]
        if model not in data:
            data[model] = {dim: [] for dim in dim_labels}
        for i, dim in enumerate(dim_labels):
            data[model][dim].append(row[i+3])

    fig, ax = plt.subplots(subplot_kw=dict(polar=True))
    labels = dim_labels
    angles = [n / float(len(labels)) * 2 * 3.14159 for n in range(len(labels))]
    angles += angles[:1]

    for model, dim_scores in data.items():
        scores = [sum(dim_scores[dim])/len(dim_scores[dim]) for dim in labels]
        scores += scores[:1]
        ax.plot(angles, scores, label=model)
        ax.fill(angles, scores, alpha=0.1)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels)
    ax.set_title("口腔模型五维评分雷达图")
    ax.legend(loc='upper right')
    plt.tight_layout()
    plt.savefig("radar_scores.png")
    plt.show()


# 示例调用（替换模型函数）
if __name__ == '__main__':
    from inference_wrappers import deepseek_inference, model_judge
    evaluate_subjective_cases("./disease_history", deepseek_inference, model_judge, "model_scores.xlsx")
    draw_radar_chart("model_scores.xlsx")
